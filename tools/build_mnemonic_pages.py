# -*- coding: utf-8 -*-
"""Build 10 new subject pages for the IPTE repo from 암기장.xlsm,
matching the existing site style (sidebar / hero / mnemonic-book /
topic-cards / flashcards)."""
import re, os, html, warnings, shutil
warnings.filterwarnings('ignore')
from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock

import sys
XLSM = sys.argv[1] if len(sys.argv) > 1 else os.path.join(os.path.dirname(__file__), '..', '암기장.xlsm')
WB = load_workbook(XLSM, keep_vba=True, rich_text=True)
REPO = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))

SUBJECTS = [
    # (sheet, file_stem, page_title, nav_title_html, hero_sub)
    ('법제도',  '01-법제도',   '법제도',            '법제도<br>기술사 정리',       '법률 · 인증제도 · 가이드라인'),
    ('신기술',  '02-신기술',   '신기술',            '신기술<br>기술사 정리',       '최신 기술 트렌드 · 가트너 전략기술'),
    ('보안',    '03-보안',     '정보보안',          '정보보안<br>기술사 정리',     '보안 개요 · 암호화 · 네트워크/시스템 보안'),
    ('빅데',    '04-빅데이터', '빅데이터 분석',      '빅데이터<br>기술사 정리',     '분석 기획 · 거버넌스 · 기법'),
    ('프관',    '05-프로젝트관리', '프로젝트 관리',  '프로젝트관리<br>기술사 정리', 'PMBOK · ISO 21500 · 관리 지식영역'),
    ('NW',      '06-네트워크', '네트워크',          '네트워크<br>기술사 정리',     'OSI · 프로토콜 · 접속장치'),
    ('CAOS',    '07-caos',     '컴퓨터구조/운영체제', 'CA·OS<br>기술사 정리',      '운영체제 · 컴퓨터 아키텍처'),
    ('알고리즘','08-알고리즘', '자료구조/알고리즘',  '알고리즘<br>기술사 정리',     '자료구조 · 알고리즘 분류'),
    ('디패',    '09-uml디자인패턴', 'UML/디자인패턴', 'UML·디자인패턴<br>기술사 정리', '객체지향 · GoF 패턴'),
    ('IT뉴스',  '10-it뉴스',   'IT 뉴스 토픽',       'IT뉴스<br>기술사 정리',      '최신 이슈 · 추가 토픽'),
]

# ---------------- rich text helpers ----------------
def runs_from_cell(value):
    if value is None:
        return '', []
    if isinstance(value, CellRichText):
        text, runs, pos = '', [], 0
        for el in value:
            if isinstance(el, TextBlock):
                t = el.text or ''
                bold = bool(el.font.b) if el.font else False
                color = None
                if el.font and el.font.color and el.font.color.type == 'rgb' and isinstance(el.font.color.rgb, str):
                    color = el.font.color.rgb[-6:]
            else:
                t, bold, color = str(el), False, None
            runs.append((pos, pos + len(t), bold, color))
            text += t
            pos += len(t)
        return text, runs
    text = str(value)
    return text, [(0, len(text), False, None)]

def esc(s):
    return html.escape(str(s)) if s else ''

def render_range(text, runs, start, end):
    out = []
    for (rs, re_, bold, color) in runs:
        if re_ <= start or rs >= end:
            continue
        seg = text[max(rs, start):min(re_, end)]
        if not seg:
            continue
        e = esc(seg)
        styles = []
        if bold:
            styles.append('font-weight:700')
        if color and color.upper() not in ('000000', 'FFFFFF'):
            styles.append(f'color:#{color}')
        out.append(f'<span style="{";".join(styles)}">{e}</span>' if styles else e)
    return ''.join(out)

def parse_title(title):
    stars = title.count('★')
    pm = re.search(r'\(p\.?\s*(\d+)\)', title)
    page = pm.group(1) if pm else None
    clean = re.sub(r'\(p\.?\s*\d+\)', '', title.replace('★', '')).strip()
    return clean, stars, page

MN_RE = re.compile(r'^[가-힣0-9()\s·,~+↑↓]{2,18}\.?$')
LEGAL_RE = re.compile(r'(법|시행령|조례)\s*(제?\s*\d+\s*조)?')

def find_mnemonics(text, runs):
    """bold black runs that look like 두음 phrases"""
    found = []
    for (rs, re_, bold, color) in runs:
        if not bold:
            continue
        if color and color.upper() not in ('000000',):
            continue
        seg = text[rs:re_].strip()
        core = seg.rstrip('.').strip()
        if not core or len(core) > 18:
            continue
        hangul = sum(1 for ch in core if '가' <= ch <= '힣')
        if hangul < 2:
            continue
        if LEGAL_RE.search(core) and '조' in core:
            continue
        if not MN_RE.match(core if not core.endswith('.') else core):
            continue
        # context: preceding [tag]
        before = text[:rs]
        tag_m = None
        for m in re.finditer(r'\[([^\]\n]{1,30})\]', before):
            tag_m = m.group(1).strip()
        # following expansion (text right after the run, up to newline or next [)
        after = text[re_:]
        exp = re.split(r'[\n\[]', after, 1)[0].strip().lstrip('.').strip()
        found.append({'key': core, 'tag': tag_m, 'expansion': exp[:120]})
    return found

def body_to_parts(value):
    """intro_html, [(tag_html, content_html, tag_plain, content_plain)]"""
    text, runs = runs_from_cell(value)
    if not text.strip():
        return '', [], text, runs
    ms = list(re.finditer(r'\[[^\]\n]{1,40}\]', text))
    if not ms:
        return render_range(text, runs, 0, len(text)).strip(), [], text, runs
    intro = render_range(text, runs, 0, ms[0].start()).strip()
    rows = []
    for i, m in enumerate(ms):
        tag_h = render_range(text, runs, m.start()+1, m.end()-1).strip()
        cs, ce = m.end(), (ms[i+1].start() if i+1 < len(ms) else len(text))
        content_h = render_range(text, runs, cs, ce).strip()
        tag_p = text[m.start()+1:m.end()-1].strip()
        content_p = re.sub(r'\s+', ' ', text[cs:ce]).strip()
        if tag_h or content_h:
            rows.append((tag_h, content_h, tag_p, content_p))
    return intro, rows, text, runs

def build_tree(rows):
    root = {'children': []}
    stack = [root]
    for row in rows:
        node = {'item': row, 'children': []}
        while len(stack) > row['indent'] + 1:
            stack.pop()
        stack[-1]['children'].append(node)
        stack.append(node)
    return root['children']

def slugify(s, fallback):
    s = re.sub(r'[^0-9A-Za-z가-힣]+', '-', s).strip('-')
    return s[:40] if s else fallback

# ---------------- page assembly ----------------
def gen_subject(sheet, stem, page_title, nav_title, hero_sub):
    ws = WB[sheet]
    rows = []
    for r in range(1, ws.max_row + 1):
        a, b = ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value
        if a is None and b is None:
            continue
        plain, _ = runs_from_cell(a)
        indent = (len(plain) - len(plain.lstrip(' '))) // 4 if plain else 0
        rows.append({'row': r, 'indent': indent, 'title': (plain or '').strip(), 'body': b})

    # drop the "<과목>" banner row and 기출문제 stubs at depth 0 with no body
    rows = [x for x in rows if not (x['title'].startswith('<') and x['title'].endswith('>'))]

    # flat glossary sheets (e.g. IT뉴스): every row is depth-0 → group under one section
    n_depth0 = sum(1 for x in rows if x['indent'] == 0)
    if rows and n_depth0 > len(rows) * 0.8 and n_depth0 > 30:
        for x in rows:
            x['indent'] += 1
        rows.insert(0, {'row': 0, 'indent': 0, 'title': '최신 토픽 용어집', 'body': None})

    # images -> repo assets
    img_dir_rel = f'assets/img/{stem}'
    img_dir = os.path.join(REPO, img_dir_rel)
    os.makedirs(img_dir, exist_ok=True)
    images_by_row = {}
    for idx, im in enumerate(ws._images):
        rownum = im.anchor._from.row + 1
        ext = im.format.lower() if getattr(im, 'format', None) else 'png'
        fname = f'img_{idx:03d}.{ext}'
        with open(os.path.join(img_dir, fname), 'wb') as f:
            f.write(im._data())
        images_by_row.setdefault(rownum, []).append(f'../{img_dir_rel}/{fname}')

    tree = build_tree(rows)

    sections, mnemonics, flashcards = [], [], []

    def topic_card(node, sec_id, idx_in_sec, sec_title):
        row = node['item']
        title, stars, page = parse_title(row['title'])
        intro_h, tagrows, text, runs = body_to_parts(row['body'])
        mns = find_mnemonics(text, runs)
        for mn in mns:
            mnemonics.append({'topic': title, 'key': mn['key'], 'tag': mn['tag'],
                              'expansion': mn['expansion'], 'sec': sec_title})
        # flashcards from tag rows (limit 2 per topic to keep deck focused)
        for (tg_h, ct_h, tg_p, ct_p) in tagrows[:2]:
            if tg_p and ct_p and len(ct_p) > 3:
                flashcards.append({'cat': sec_title[:12], 'q': f'{title} — [{tg_p}]?',
                                   'a': ct_p[:220]})
        star_badge = f'<span class="exam-tag">{"★"*stars} 빈출</span>' if stars else ''
        page_p = f'<p class="pages">교재 p.{page}</p>' if page else ''
        mn_pill = f'<span class="mnemonic">암기: {esc(mns[0]["key"])}</span>' if mns else ''
        chips = ''
        if tagrows:
            chips = '<div class="chips">' + ''.join(f'<span>{esc(t[2])}</span>' for t in tagrows[:8]) + '</div>'
        table = ''
        if tagrows:
            body_rows = ''.join(
                f'<tr><td>{tg}</td><td>{ct}</td></tr>' for tg, ct, _, _ in tagrows)
            table = (f'<div class="table-wrap"><table><thead><tr><th>구분</th><th>내용</th></tr></thead>'
                     f'<tbody>{body_rows}</tbody></table></div>')
        intro_p = f'<p>{star_badge}{intro_h}</p>' if intro_h else (f'<p>{star_badge}</p>' if star_badge else '')
        figs = ''.join(
            f'<figure style="margin:14px 0"><img src="{src}" alt="{esc(title)}" loading="lazy" '
            f'style="width:100%;height:auto;border:1px solid var(--line);border-radius:12px"></figure>'
            for src in images_by_row.get(row['row'], []))
        # child topics rendered as collapsible depth block
        child_html = ''
        if node['children']:
            inner = ''.join(topic_card(c, sec_id, None, sec_title) for c in node['children'])
            child_html = (f'<details class="topic-depth-wrap"><summary>'
                          f'<span class="depth-title">📂 하위 토픽 ({len(node["children"])})</span>'
                          f'<span class="depth-hint">접기/펼치기</span></summary>'
                          f'<div class="topic-depth-body">{inner}</div></details>')
        card_id = f' id="{sec_id}-{idx_in_sec}"' if idx_in_sec is not None else ''
        return (f'<article{card_id} class="topic-card">'
                f'<div class="topic-head"><div><h3>{esc(title)}</h3>{page_p}</div>{mn_pill}</div>'
                f'{chips}{intro_p}{table}{figs}{child_html}</article>')

    for si, top in enumerate(tree):
        row = top['item']
        title, stars, page = parse_title(row['title'])
        if not title:
            continue
        sec_id = 's' + str(si) + '-' + slugify(title, 'sec')
        # a depth-0 node may itself have body -> render as its own card first
        cards = ''
        if row['body'] or images_by_row.get(row['row']):
            cards += topic_card({'item': row, 'children': []}, sec_id, 0, title)
        for ci, child in enumerate(top['children'], start=1):
            cards += topic_card(child, sec_id, ci, title)
        if cards:
            sections.append({'id': sec_id, 'title': title, 'html': cards})

    # ---- assemble page ----
    nav_items = ['<a href="#mnemonic-book">📖 두음 모음집</a>']
    for i, s in enumerate(sections, 1):
        nav_items.append(f'<a href="#{s["id"]}">{i}. {esc(s["title"])}</a>')
    nav_items.append('<a href="#flashcards">암기카드</a>')
    nav_html = ''.join(nav_items)

    mn_items = ''
    for i, mn in enumerate(mnemonics, 1):
        src = f'{i}. {esc(mn["topic"])}' + (f' · {esc(mn["tag"])}' if mn['tag'] else '')
        more = (f'<details class="mn-more"><summary>설명 보기</summary>'
                f'<p>{esc(mn["expansion"])}</p></details>') if mn['expansion'] else ''
        mn_items += (f'<details class="mnemonic-item"><summary>'
                     f'<span class="mn-from">{src}</span><span class="mn-key">{esc(mn["key"])}</span></summary>'
                     f'<div class="mn-body"><p class="mn-split"><b>출처:</b> {esc(mn["sec"])}</p>{more}</div></details>')
    mn_section = (f'<section class="panel mnemonic-book-section" id="mnemonic-book">'
                  f'<details class="mnemonic-book-wrap"><summary>'
                  f'<span class="mn-wrap-title">📖 두음 모음집 — {esc(page_title)}</span>'
                  f'<span class="mn-wrap-hint">전체 접기/펼치기</span></summary>'
                  f'<div class="mnemonic-book-body">'
                  f'<p class="subtle">암기장 기준 · 출처 → 두음 → 설명 ({len(mnemonics)}개)</p>'
                  f'<input class="mn-filter" type="search" placeholder="두음 검색…" oninput="filterMnemonics(this)"/>'
                  f'<div class="mnemonic-list">{mn_items}</div></div></details></section>') if mnemonics else ''

    sec_html = ''
    for i, s in enumerate(sections, 1):
        sec_html += f'<section id="{s["id"]}" class="section"><h2>{i}. {esc(s["title"])}</h2>{s["html"]}</section>'

    cards_html = ''.join(
        f'<div class="flash-card" data-cat="{esc(fc["cat"])}">'
        f'<div class="card-front"><span>{esc(fc["cat"])}</span><p>{esc(fc["q"])}</p></div>'
        f'<div class="card-back"><pre>{esc(fc["a"])}</pre></div></div>'
        for fc in flashcards)
    flash_section = (f'<section id="flashcards" class="flash-section"><h2>암기카드</h2>'
                     f'<div class="flash-controls">'
                     f'<input id="search" type="search" placeholder="검색…"/>'
                     f'<select id="catFilter"><option value="">전체 분류</option></select>'
                     f'<button id="shuffleBtn" type="button">섞기</button>'
                     f'<button id="revealBtn" type="button">전체 뒤집기</button>'
                     f'<button id="resetBtn" type="button">초기화</button></div>'
                     f'<div class="flash-grid" id="cards">{cards_html}</div></section>') if flashcards else ''

    js = """<script>
const cards = Array.from(document.querySelectorAll('.flash-card'));
cards.forEach(c => c.addEventListener('click', () => c.classList.toggle('flipped')));
const catFilter = document.getElementById('catFilter');
if (catFilter) {
  [...new Set(cards.map(c => c.dataset.cat))].sort().forEach(cat => {
    const opt = document.createElement('option'); opt.value = cat; opt.textContent = cat; catFilter.appendChild(opt);
  });
  function applyFilter() {
    const q = (document.getElementById('search').value || '').toLowerCase();
    const cat = catFilter.value;
    cards.forEach(c => {
      const text = c.textContent.toLowerCase();
      const show = (!q || text.includes(q)) && (!cat || c.dataset.cat === cat);
      c.style.display = show ? '' : 'none';
    });
  }
  document.getElementById('search').addEventListener('input', applyFilter);
  catFilter.addEventListener('change', applyFilter);
  document.getElementById('shuffleBtn').addEventListener('click', () => {
    const wrap = document.getElementById('cards');
    cards.sort(() => Math.random() - 0.5).forEach(c => wrap.appendChild(c));
  });
  document.getElementById('revealBtn').addEventListener('click', () => cards.forEach(c => c.classList.add('flipped')));
  document.getElementById('resetBtn').addEventListener('click', () => {
    document.getElementById('search').value=''; catFilter.value=''; cards.forEach(c => { c.style.display=''; c.classList.remove('flipped'); });
  });
}
function filterMnemonics(inp){
  const q=(inp.value||'').toLowerCase().trim();
  const list=inp.closest('.mnemonic-book-section');
  if(!list) return;
  list.querySelectorAll('.mnemonic-item').forEach(el=>{
    el.classList.toggle('hidden', q && !el.textContent.toLowerCase().includes(q));
  });
}
const navLinks = Array.from(document.querySelectorAll('.sidebar nav a'));
if (navLinks.length) {
  const obs = new IntersectionObserver(entries => {
    entries.forEach(e => {
      if (e.isIntersecting) {
        navLinks.forEach(l => l.classList.toggle('active', l.getAttribute('href') === '#' + e.target.id));
      }
    });
  }, { rootMargin: '-35% 0px -60% 0px', threshold: 0 });
  document.querySelectorAll('section[id], article[id]').forEach(s => obs.observe(s));
}
</script>"""

    page = f"""<!doctype html>
<html lang="ko">
<head>
<meta charset="utf-8" />
<meta name="viewport" content="width=device-width, initial-scale=1" />
<title>{esc(page_title)} 기술사 분석 키워드 &amp; 암기카드</title>
<link rel="stylesheet" href="../assets/study-page.css"/>
<link rel="stylesheet" href="../assets/diagrams.css"/>
<link rel="stylesheet" href="../assets/mnemonic-book.css"/>
</head>
<body>
<aside class="sidebar">
  <a class="back-home" href="../index.html">← 목차로</a>
  <h1>{nav_title}</h1>
  <p class="sub">키워드 + 도식 + 암기카드</p>
  <nav>{nav_html}</nav>
</aside>
<main class="main">
  <a class="back-home back-home-main" href="../index.html">← 목차로</a>
  <details class="mobile-toc"><summary>목차 열기</summary><div class="toc-links">{nav_html}</div></details>
  <div class="hero">
    <h1>{esc(page_title)} 기술사 분석 키워드 &amp; 암기카드</h1>
    <p>{esc(hero_sub)}</p>
    <p class="note">PC: 왼쪽 고정 목차 / 모바일: 상단 접기형 목차 / 표: 가로 스크롤 지원</p>
  </div>
{mn_section}
{sec_html}
{flash_section}
{js}
</body>
</html>
"""
    out = os.path.join(REPO, 'pages', f'{stem}.html')
    with open(out, 'w', encoding='utf-8') as f:
        f.write(page)
    n_imgs = sum(len(v) for v in images_by_row.values())
    print(f'{sheet:6s} -> pages/{stem}.html  sections={len(sections)} mnemonics={len(mnemonics)} cards={len(flashcards)} imgs={n_imgs}')

for args in SUBJECTS:
    gen_subject(*args)
